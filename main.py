"""
Data Cleaner Script
===================
Bu Python aracı, verilen bir CSV dosyasını temizleyip normalize eder.
İşlemler:
- Eksik (NaN) verileri kaldırır
- Tekrarlanan (duplicate) satırları siler
- Metin verilerini temizler (boşlukları kaldırır, küçük harfe çevirir)
- Sayısal verileri normalize eder (MinMaxScaler ile 0–1 aralığına)
- Sonucu 'Temiz_Veri.xlsx' olarak kaydeder

Kullanım:
---------
python main.py
"""

import pandas as pd
import numpy as np
import openpyxl
from sklearn.preprocessing import MinMaxScaler


def clean_csv_file(filename: str, output_filename: str = "Temiz_Veri.xlsx"):
    """CSV dosyasını temizleyip Excel olarak kaydeder."""

    # CSV dosyasını oku
    data = pd.read_csv(filename)

    print("\n--- ORİJİNAL VERİ ---")
    print("Toplam satır:", len(data))
    print("Toplam sütun:", len(data.columns))

    # Eksik değerleri sil
    data = data.dropna()

    # Tekrarlanan satırları sil
    data = data.drop_duplicates()

    # Metin kolonlarını düzenle
    for col in data.select_dtypes(include=['object']).columns:
        data[col] = data[col].str.strip().str.lower()

    # Sayısal kolonları normalize et
    num_cols = data.select_dtypes(include=np.number).columns
    if len(num_cols) > 0:
        scaler = MinMaxScaler()
        data[num_cols] = scaler.fit_transform(data[num_cols])

    print("\n--- TEMİZLENMİŞ VERİ ---")
    print("Toplam satır:", len(data))
    print("Toplam sütun:", len(data.columns))
    print("NaN silindi, duplicate silindi, metinler düzenlendi, sayısal kolonlar normalize edildi.\n")

    # Excel'e kaydet
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Başlıkları ekle
    for x, col in enumerate(data.columns, start=1):
        sheet.cell(row=1, column=x, value=col)

    # Verileri ekle
    for i in range(len(data)):
        for j, col in enumerate(data.columns, start=1):
            sheet.cell(row=i + 2, column=j, value=data.iat[i, j])

    wb.save(output_filename)
    print(f"✅ İşlem başarıyla tamamlandı. '{output_filename}' oluşturuldu.")


if __name__ == "__main__":
    dosya = input("Temizlenecek CSV dosyasının adını girin (örnek: data.csv): ").strip()
    clean_csv_file(dosya)
